import os
import pytest
import pandas as pd
from openpyxl import load_workbook
from reporting import ReportGenerator


@pytest.fixture
def rg():
    return ReportGenerator(output_format="excel")


@pytest.fixture
def rg_csv():
    return ReportGenerator(output_format="csv")


@pytest.fixture(autouse=True)
def cleanup(tmp_path, monkeypatch):
    """Run all tests from a temp directory so generated files are auto-cleaned."""
    monkeypatch.chdir(tmp_path)


# --- P&L Export ---

class TestPLExport:
    def test_simple_pl_excel(self, rg):
        pl = {"Total Revenue": 10000, "Total Expenses": 4000, "Net Income": 6000}
        rg.export_profit_loss(pl, file_name="test_pl")
        assert os.path.exists("test_pl.xlsx")

        wb = load_workbook("test_pl.xlsx")
        ws = wb.active
        assert ws["A1"].value == "Profit & Loss Statement"
        assert ws["A2"].value == "Period: All Dates"

    def test_pl_with_dates(self, rg):
        pl = {"Total Revenue": 10000, "Total Expenses": 4000, "Net Income": 6000}
        rg.export_profit_loss(pl, file_name="test_pl", start_date="2025-01-01", end_date="2025-12-31")
        wb = load_workbook("test_pl.xlsx")
        ws = wb.active
        assert "2025-01-01" in ws["A2"].value
        assert "2025-12-31" in ws["A2"].value

    def test_detailed_pl_has_sections(self, rg):
        pl = {
            "revenue_detail": {"Consulting": 8000, "Products": 2000},
            "expense_detail": {"Rent": 3000, "Payroll": 1000},
            "Total Revenue": 10000,
            "Total Expenses": 4000,
            "Gross Profit": 6000,
            "Net Income": 6000,
        }
        rg.export_profit_loss(pl, file_name="test_pl_detail")
        wb = load_workbook("test_pl_detail.xlsx")
        ws = wb.active

        # Collect all values in column A
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Revenue" in values
        assert "Expenses" in values
        assert "Net Income" in values
        assert "Gross Profit" in values

    def test_pl_csv(self, rg_csv):
        pl = {"Total Revenue": 10000, "Total Expenses": 4000, "Net Income": 6000}
        rg_csv.export_profit_loss(pl, file_name="test_pl")
        assert os.path.exists("test_pl.csv")
        df = pd.read_csv("test_pl.csv")
        assert "Net Income" in df["Metric"].values

    def test_net_income_highlighted(self, rg):
        pl = {"Total Revenue": 10000, "Total Expenses": 4000, "Net Income": 6000}
        rg.export_profit_loss(pl, file_name="test_pl")
        wb = load_workbook("test_pl.xlsx")
        ws = wb.active
        # Find the Net Income row
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value == "Net Income":
                assert cell.font.bold
                break


# --- Balance Sheet Export ---

class TestBSExport:
    def test_flat_dict_excel(self, rg):
        bs = {"Cash": 5000, "Loan": 2000}
        rg.export_balance_sheet(bs, file_name="test_bs")
        assert os.path.exists("test_bs.xlsx")
        wb = load_workbook("test_bs.xlsx")
        ws = wb.active
        assert ws["A1"].value == "Balance Sheet"

    def test_grouped_excel(self, rg):
        bs = {
            "rows": [
                {"Account": "Bank", "Balance": None, "row_type": "section"},
                {"Account": "  Checking", "Balance": 10000, "row_type": "detail"},
                {"Account": "Total Bank", "Balance": 10000, "row_type": "subtotal"},
                {"Account": "", "Balance": None, "row_type": "spacer"},
            ],
            "section_totals": {"Assets": 10000},
        }
        rg.export_balance_sheet(bs, file_name="test_bs_grp", report_date="2025-12-31")
        wb = load_workbook("test_bs_grp.xlsx")
        ws = wb.active
        assert "2025-12-31" in ws["A2"].value

        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Total Assets" in values

    def test_bs_csv(self, rg_csv):
        bs = {"Cash": 5000, "Loan": 2000}
        rg_csv.export_balance_sheet(bs, file_name="test_bs")
        assert os.path.exists("test_bs.csv")

    def test_dataframe_input(self, rg):
        df = pd.DataFrame({"Account": ["Cash"], "Balance": [1000]})
        rg.export_balance_sheet(df, file_name="test_bs_df")
        assert os.path.exists("test_bs_df.xlsx")

    def test_currency_format(self, rg):
        bs = {"Cash": 5000.50}
        rg.export_balance_sheet(bs, file_name="test_bs")
        wb = load_workbook("test_bs.xlsx")
        ws = wb.active
        # Check that Balance column has currency format
        for row in ws.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
            assert "$" in row[0].number_format


# --- Cash Flow Export ---

class TestCFExport:
    def _sample_cf(self):
        return {
            "net_income": 10000,
            "operating": {"Accounts Receivable": -3000, "Accounts Payable": 2000},
            "net_operating": 9000,
            "investing": {"Fixed Assets": -5000},
            "net_investing": -5000,
            "financing": {"Equity": 20000},
            "net_financing": 20000,
            "net_change_in_cash": 24000,
            "ending_cash": 15000,
        }

    def test_cf_excel_creates_file(self, rg):
        rg.export_cash_flow(self._sample_cf(), file_name="test_cf")
        assert os.path.exists("test_cf.xlsx")

    def test_cf_excel_title(self, rg):
        rg.export_cash_flow(self._sample_cf(), file_name="test_cf")
        wb = load_workbook("test_cf.xlsx")
        ws = wb.active
        assert ws["A1"].value == "Cash Flow Statement"

    def test_cf_excel_sections(self, rg):
        rg.export_cash_flow(self._sample_cf(), file_name="test_cf")
        wb = load_workbook("test_cf.xlsx")
        ws = wb.active
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Cash Flows from Operating Activities" in values
        assert "Cash Flows from Investing Activities" in values
        assert "Cash Flows from Financing Activities" in values
        assert "Net Change in Cash" in values
        assert "Ending Cash Balance" in values

    def test_cf_excel_with_dates(self, rg):
        rg.export_cash_flow(self._sample_cf(), file_name="test_cf",
                            start_date="2025-01-01", end_date="2025-12-31")
        wb = load_workbook("test_cf.xlsx")
        ws = wb.active
        assert "2025-01-01" in ws["A2"].value

    def test_cf_csv(self, rg_csv):
        rg_csv.export_cash_flow(self._sample_cf(), file_name="test_cf")
        assert os.path.exists("test_cf.csv")
        df = pd.read_csv("test_cf.csv")
        assert "Net Change in Cash" in df["Item"].values
        assert "Ending Cash Balance" in df["Item"].values

    def test_cf_ending_cash_highlighted(self, rg):
        rg.export_cash_flow(self._sample_cf(), file_name="test_cf")
        wb = load_workbook("test_cf.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value == "Ending Cash Balance":
                assert cell.font.bold
                break
