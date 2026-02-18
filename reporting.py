# reporting.py
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """
    Handles final report formatting and exporting (Excel, CSV, etc.).
    """

    def __init__(self, output_format="excel"):
        """
        :param output_format: 'excel' or 'csv'
        """
        self.output_format = output_format.lower()

        # Shared styles
        self._header_font = Font(bold=True, size=11, color="FFFFFF")
        self._header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self._title_font = Font(bold=True, size=14, color="1F3864")
        self._subtitle_font = Font(size=11, color="404040", italic=True)
        self._currency_format = '$#,##0.00'
        self._thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        self._highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self._highlight_font = Font(bold=True, size=11)
        self._section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        self._section_font = Font(bold=True, size=11, color="1F3864")

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _build_pl_rows(self, pl_statement):
        """Build row data for the P&L report from either detailed or simple format."""
        rows = []  # list of (label, amount, style) where style is 'section', 'detail', 'subtotal', or 'total'

        has_detail = "revenue_detail" in pl_statement

        if has_detail:
            # Revenue section
            rows.append(("Revenue", None, "section"))
            for cat, amt in pl_statement["revenue_detail"].items():
                rows.append((f"  {cat}", amt, "detail"))
            rows.append(("Total Revenue", pl_statement["Total Revenue"], "subtotal"))

            rows.append(("", None, "spacer"))

            # Expense section
            rows.append(("Expenses", None, "section"))
            for cat, amt in pl_statement["expense_detail"].items():
                rows.append((f"  {cat}", amt, "detail"))
            rows.append(("Total Expenses", pl_statement["Total Expenses"], "subtotal"))

            rows.append(("", None, "spacer"))

            # Gross Profit
            rows.append(("Gross Profit", pl_statement.get("Gross Profit", 0), "subtotal"))

            rows.append(("", None, "spacer"))
        else:
            rows.append(("Total Revenue", pl_statement.get("Total Revenue", 0), "detail"))
            rows.append(("Total Expenses", pl_statement.get("Total Expenses", 0), "detail"))
            rows.append(("", None, "spacer"))

        # Net Income (always last)
        rows.append(("Net Income", pl_statement.get("Net Income", 0), "total"))

        return rows

    def export_profit_loss(self, pl_statement, file_name="profit_loss_report",
                           start_date=None, end_date=None):
        """
        Save a P&L statement in Excel or CSV format.
        Supports both simple (3-line) and detailed (categorized) formats.
        """
        rows = self._build_pl_rows(pl_statement)

        if self.output_format == "csv":
            df = pd.DataFrame([(r[0], r[1]) for r in rows if r[2] != "spacer"],
                              columns=["Metric", "Amount"])
            df.to_csv(f"{file_name}.csv", index=False)
            print(f"[INFO] P&L report saved as: {file_name}.csv")
            return

        path = f"{file_name}.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"

        # Title
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = "Profit & Loss Statement"
        title_cell.font = self._title_font
        title_cell.alignment = Alignment(horizontal="left")

        # Subtitle
        ws.merge_cells("A2:B2")
        subtitle_cell = ws["A2"]
        if start_date and end_date:
            subtitle_cell.value = f"Period: {start_date} to {end_date}"
        else:
            subtitle_cell.value = "Period: All Dates"
        subtitle_cell.font = self._subtitle_font

        # Column headers on row 4
        for col_num, header in enumerate(["Metric", "Amount"], 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._thin_border

        # Write data rows starting at row 5
        current_row = 5
        for label, amount, style in rows:
            if style == "spacer":
                current_row += 1
                continue

            label_cell = ws.cell(row=current_row, column=1, value=label)
            amt_cell = ws.cell(row=current_row, column=2, value=amount)

            # Apply borders and currency to all non-spacer rows
            for c in (label_cell, amt_cell):
                c.border = self._thin_border
            if amount is not None:
                amt_cell.number_format = self._currency_format
                amt_cell.alignment = Alignment(horizontal="right")

            if style == "section":
                label_cell.font = self._section_font
                label_cell.fill = self._section_fill
                amt_cell.fill = self._section_fill
            elif style == "subtotal":
                label_cell.font = Font(bold=True, size=11)
                amt_cell.font = Font(bold=True, size=11)
            elif style == "total":
                label_cell.font = self._highlight_font
                label_cell.fill = self._highlight_fill
                amt_cell.font = self._highlight_font
                amt_cell.fill = self._highlight_fill

            current_row += 1

        self._auto_fit_columns(ws)
        wb.save(path)
        print(f"[INFO] P&L report saved as: {path}")

    def export_balance_sheet(self, balance_data, file_name="balance_sheet_report",
                             report_date=None):
        """
        Export a Balance Sheet.
        Accepts structured data (with rows + section_totals) or a flat dict/DataFrame.
        """
        # Determine if we have the new grouped format
        is_grouped = isinstance(balance_data, dict) and "rows" in balance_data

        if not is_grouped:
            if isinstance(balance_data, dict):
                df = pd.DataFrame(list(balance_data.items()), columns=["Account", "Balance"])
            else:
                df = balance_data
            return self._export_flat_balance_sheet(df, file_name, report_date)

        rows = balance_data["rows"]
        section_totals = balance_data.get("section_totals", {})

        if self.output_format == "csv":
            csv_rows = [(r["Account"], r["Balance"]) for r in rows if r["row_type"] != "spacer"]
            df = pd.DataFrame(csv_rows, columns=["Account", "Balance"])
            df.to_csv(f"{file_name}.csv", index=False)
            print(f"[INFO] Balance Sheet report saved as: {file_name}.csv")
            return

        path = f"{file_name}.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Title
        ws.merge_cells("A1:B1")
        ws["A1"].value = "Balance Sheet"
        ws["A1"].font = self._title_font
        ws["A1"].alignment = Alignment(horizontal="left")

        # Subtitle
        ws.merge_cells("A2:B2")
        ws["A2"].value = f"As of: {report_date}" if report_date else "As of: Current"
        ws["A2"].font = self._subtitle_font

        # Column headers
        for col_num, header in enumerate(["Account", "Balance"], 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._thin_border

        # Write rows
        current_row = 5
        for row in rows:
            rt = row["row_type"]
            if rt == "spacer":
                current_row += 1
                continue

            acct_cell = ws.cell(row=current_row, column=1, value=row["Account"])
            bal_cell = ws.cell(row=current_row, column=2, value=row["Balance"])

            for c in (acct_cell, bal_cell):
                c.border = self._thin_border
            if row["Balance"] is not None:
                bal_cell.number_format = self._currency_format
                bal_cell.alignment = Alignment(horizontal="right")

            if rt == "section":
                acct_cell.font = self._section_font
                acct_cell.fill = self._section_fill
                bal_cell.fill = self._section_fill
            elif rt == "subtotal":
                acct_cell.font = Font(bold=True, size=11)
                bal_cell.font = Font(bold=True, size=11)

            current_row += 1

        # Summary: Total Assets / Liabilities / Equity
        current_row += 1
        for group in ["Assets", "Liabilities", "Equity"]:
            if group not in section_totals:
                continue
            acct_cell = ws.cell(row=current_row, column=1, value=f"Total {group}")
            bal_cell = ws.cell(row=current_row, column=2, value=section_totals[group])
            acct_cell.font = self._highlight_font
            acct_cell.fill = self._highlight_fill
            bal_cell.font = self._highlight_font
            bal_cell.fill = self._highlight_fill
            bal_cell.number_format = self._currency_format
            bal_cell.alignment = Alignment(horizontal="right")
            for c in (acct_cell, bal_cell):
                c.border = self._thin_border
            current_row += 1

        self._auto_fit_columns(ws)
        wb.save(path)
        print(f"[INFO] Balance Sheet report saved as: {path}")

    def _export_flat_balance_sheet(self, df, file_name, report_date):
        """Flat balance sheet export (no account grouping)."""
        if self.output_format == "csv":
            df.to_csv(f"{file_name}.csv", index=False)
            print(f"[INFO] Balance Sheet report saved as: {file_name}.csv")
            return

        path = f"{file_name}.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Balance Sheet", startrow=3)
            ws = writer.sheets["Balance Sheet"]
            num_cols = len(df.columns)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            ws["A1"].value = "Balance Sheet"
            ws["A1"].font = self._title_font
            ws["A1"].alignment = Alignment(horizontal="left")

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            ws["A2"].value = f"As of: {report_date}" if report_date else "As of: Current"
            ws["A2"].font = self._subtitle_font

            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=4, column=col_num)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = self._thin_border

            for row_num in range(5, 5 + len(df)):
                for col_num in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = self._thin_border
                    if col_num == num_cols:
                        cell.number_format = self._currency_format
                        cell.alignment = Alignment(horizontal="right")

            self._auto_fit_columns(ws)

        print(f"[INFO] Balance Sheet report saved as: {path}")
